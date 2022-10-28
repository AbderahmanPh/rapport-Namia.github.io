import pandas as pd
from pandas import ExcelWriter
import numpy as np
import os
from openpyxl import load_workbook

###---------------------------------------- cleaning bah_files --------------------------------------------------------
## sorting months reference
months = {f'{k.upper()}.xlsx':v for (k,v) in zip(["janvier","fevrier","mars","avril","mai","juin","juillet","aout","septembre","octobre","novembre","decembre"], [1,2,3,4,5,6,7,8,9,10,11,12]) }

months_31 = [f'{x.upper()}.xlsx' for x in ['janvier', 'mars', 'mai', 'juillet', 'aout', 'octobre', 'decembre']]
months_30 = [f'{x.upper()}.xlsx' for x in ['avril', 'juin', 'septembre', 'novembre']]

for file in [month for month in sorted([file for file in os.listdir('./bah_files/') if (not file.startswith('.')) ], key=lambda k: months[k] )]:
    print(file)
    
    if file in months_31:
        days = 31
    elif file in months_30:
        days = 30
    else:
        days = 28
                
    n = 0

    for i in range(days):
        ## read the sheet
        df = pd.read_excel(f'./bah_files/{file}',
                                   sheet_name=i,
                                   nrows=100)
        ## check if sheet is empty
        if df.empty:
            n += 1
            continue

        ## rename the first col
        df.rename(columns={'N° CPT': 'CODE'}, inplace=True)

        ## drop unwanted columns
        unwanted_cols = df.columns[1:12]
        df.drop(columns=unwanted_cols, axis=1, inplace=True)
        ## drop unwanted rows
        df = df[~df['CODE'].isna()]

        ## add 'date' column
        idx = 1
        date_bah_files = pd.date_range(start=f'2022-{months[file]}-01', end=f'2022-{months[file]}-{days}')[n]
        df.insert(loc=idx, column='DATE', value=date_bah_files)
        df['DATE'] = df['DATE'].dt.date

        n += 1

        ## subtotal
        cols = df.columns[2:]
        df = df.groupby(['CODE', 'DATE'],as_index=False)[cols].sum()

        ## sort by 'TOTAL'
        df.sort_values('TOTAL', ascending=False, inplace=True)

        ## save the file
        num_sheet = i + 1
        xls_path = f'./parsed_files/{file}'

        def append_df_to_excel(filename=xls_path, df=df, sheet_name=f'{num_sheet}', startrow=None,
                           truncate_sheet=False, 
                           **to_excel_kwargs):

            # Excel file doesn't exist - saving and exiting
            if not os.path.isfile(filename):
                df.to_excel(
                    filename,
                    sheet_name=sheet_name, 
                    startrow=startrow if startrow is not None else 0, 
                    **to_excel_kwargs, index=False)
                return

            # ignore [engine] parameter if it was passed
            if 'engine' in to_excel_kwargs:
                to_excel_kwargs.pop('engine')

            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

            if startrow is None:
                startrow = 0

            # write out the new sheet
            df.to_excel(writer, sheet_name, startrow=startrow, index=False)

            # save the workbook
            writer.save()

        append_df_to_excel()



## ------------------------------------- MELT AND CONCAT PARCED FILES ------------------------------------------------
## reference for sorting months
months = {f'{k.upper()}.xlsx':v for (k,v) in zip(["janvier","fevrier","mars","avril","mai","juin","juillet","aout","septembre","octobre","novembre","decembre"], [0,1,2,3,4,5,6,7,8,9,10,11]) }
frames = []

# [month for month in sorted(os.listdir('./parsed_files/'), key=lambda k: months[k]) ]

for f in [month for month in sorted([file for file in os.listdir('./parsed_files/') if (not file.startswith('.')) ], key=lambda k: months[k] )]:
    # concat all sheets in one df  
    dff = pd.concat(pd.read_excel(f"./parsed_files/{f}", sheet_name=None), ignore_index=True)
    frames.append(dff)

result_df = pd.concat(frames)
result_df.drop(columns=['Unnamed: 23'], inplace=True)

# ## melt all 'products' in one column
produit_df = result_df.melt(id_vars=['CODE', 'DATE'], 
                  value_vars=result_df.columns[:-1],
                  var_name='PRODUIT', value_name='QUANTITE')

## groupby Date
produit_by_day = produit_df.groupby(['DATE', 'PRODUIT'])[['QUANTITE']].sum().astype('int').reset_index()
produit_by_day['DATE'] = produit_by_day['DATE'].dt.date

code_by_day = produit_df.groupby(['DATE', 'CODE'])[['QUANTITE']].sum().astype('int').reset_index()
code_by_day['DATE'] = code_by_day['DATE'].dt.date

# convert 'Code' col to str for <visualization purposes>
code_by_day['CODE'] = code_by_day['CODE'].astype('str')

## save to excel file
code_by_day.to_excel('../assets/namia/code_by_day.xlsx', index=False)
produit_by_day.to_excel('../assets/namia/produit_by_day.xlsx', index=False)


