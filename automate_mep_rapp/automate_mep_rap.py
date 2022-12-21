import pandas as pd
from openpyxl import load_workbook


## READ FILES
mep = pd.read_excel('MEP.xls', index_col=False)
wb = load_workbook('rapprochement.xlsx')

### OPERATIONS
###----------------------------------------------------- MEP ----------------------------------------------------------
## MEP

## drop unuseful cols
mep.drop(columns=['Unnamed: 1', 'Unnamed: 10'], inplace=True)
## rename cols
mep.rename(columns={'Date': 'DATE', 3: 'T03', 4: 'T04', 5: 'T05', 6: 'T06', 7: 'T07', 'TR1': 'PR1', 'TR2': 'PR2', 'TR3': 'PR3'},
          inplace=True)

## drop the last row
mep.drop(mep.tail(1).index, inplace=True) # drop last n rows

## fill na with 0
mep.fillna(0, inplace=True)

## convert to int
cols = mep.columns
mep[cols[1:]] = mep[cols[1:]].astype('int')
## set a range date
mep['DATE'] = pd.date_range(start="2022-12-08",end="2022-12-18")
# .append(pd.date_range(start='2022-08-26', end='2022-09-30'))
## REMOVE 00:00:00
mep['DATE'] = mep['DATE'].dt.date

## add TOTAL column
mep['TOTAL'] = mep[cols[1:]].sum(axis=1)

## save to new file (READY FOR THE APP)
mep.to_excel('../assets/namia/mep_rapport.xlsx', index=False)


### ------------------------------------------------- Rapprochement -----------------------------------------------------
sheets = wb.sheetnames
frames = []

for sh in sheets:
    df = pd.DataFrame(data = wb[f'{sh}'].values)
    # reassign columns
    df = df.rename(columns=df.iloc[0]).drop(df.index[0])
    
    if wb[f'{sh}'].max_column > 8:
        ## STACK all columns to one col and groupby each one
        df_res = df.stack().reset_index().groupby(['level_1']).sum(numeric_only=True).T
        
        ##  order cols
        ordered_cols = df_res.columns[3:].append(df_res.columns[:3])
        df_res = df_res[ordered_cols]
        ## add 'DATE' col
        df_res.insert(0, 'DATE', f'{sh}-2022')
        frames.append(df_res)
    else:
        df = pd.DataFrame(df.sum()).T

        ## add 'DATE' col
        df.insert(0, 'DATE', f'{sh}-2022')
        frames.append(df)
    
result = pd.concat(frames, ignore_index=True)

## df to concat TRs with their relevant in the main df
df_TRS = result[['DATE', 'TR1', 'TR2', 'TR3']]
df_TRS = df_TRS.rename(columns={'TR1': 'PR1', 'TR2': 'PR2', 'TR3': 'PR3'})

# drop unwanted cols
result.drop(columns=['TR1', 'TR2', 'TR3'], inplace=True)

last_frames = [result, df_TRS]
rapprochement_df = pd.concat(last_frames, ignore_index=True)

# Drop last column of a dataframe (PR2)
# rapprochement_df = rapprochement_df.iloc[: , :-1]

## convert date to format: y-m-d
rapprochement_df['DATE'] = pd.to_datetime(rapprochement_df['DATE'], dayfirst=True)
## groupby 'DATE'
rapprochement_df = rapprochement_df.groupby('DATE').sum().reset_index()

## convert date back to string (for the search) 
rapprochement_df['DATE'] = rapprochement_df['DATE'].astype('str')

## save to new file (READY FOR THE APP)
rapprochement_df.to_excel('../assets/namia/rapprochement_rapport.xlsx', index=False)